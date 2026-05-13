# File Version: 1.0
import os
import time
import json
import re
import datetime
import urllib.request
import threading
import asyncio

from panel_globals import (
    BASE_DIR, SENSOR_CACHE_LOCK, SYSTEM_CACHE, SYSTEM_CACHE_EVENT,
    PUBLIC_STATUS_CACHE_LOCK, PUBLIC_STATUS_CACHE, PUBLIC_STATUS_FIELDS,
    SHIFT_CACHE_XLSX, SHIFT_CACHE_META_JSON, PC_PLUG_QUERY_CACHE,
    PC_PLUG_QUERY_INTERVAL_SECONDS, MEDIA_POLL_YOUTUBE_SECONDS
)
from panel_bootstrap import (
    _get_setting_bool, _get_performance_int, _get_performance_interval_seconds,
    _get_shift_share_url, _get_shift_sheet_name, _get_shift_employee_name,
    _get_shift_name_column, _get_shift_date_row
)
from panel_logging import log_error, log_hwinfo_error
from panel_state import mark_system_cache_changed
from panel_audio_controls import (
    sync_target_volume_from_system, _get_audio_refresh_interval_seconds
)
from panel_network import get_network_speed_mbps, _get_network_refresh_interval_seconds
from panel_system import get_uptime_string, _get_uptime_refresh_interval_seconds
from panel_media import update_media_and_lyrics_cache
from panel_hwinfo_snapshot import _read_latest_hwinfo_snapshot
from panel_hwinfo_snapshot import _apply_hwinfo_payload_to_cache
from panel_bootstrap import _get_hwinfo_worker_refresh_interval_seconds
from panel_tuya import _get_pc_plug_power_w, tuya_public_devices_payload

SHIFT_RUNTIME = {
    "download_date": None,
    "schedule_by_date": {},
    "last_target_iso": None,
}

def _sleep_until(next_at, minimum_sleep=0.05, fallback_sleep=0.5):
    try:
        delay = float(next_at) - time.time()
    except Exception:
        delay = float(fallback_sleep)
    time.sleep(max(float(minimum_sleep), delay if delay > 0 else float(minimum_sleep)))

def media_updater_loop():
    next_at = 0.0
    while True:
        try:
            now = time.time()
            if now >= next_at:
                next_at = _update_media_cache_tick(now)
        except Exception as e:
            log_error(f"Media updater loop error: {e}")
            next_at = time.time() + 1.0
        _sleep_until(next_at, minimum_sleep=0.05, fallback_sleep=0.6)

def volume_updater_loop():
    next_at = 0.0
    while True:
        try:
            now = time.time()
            if now >= next_at:
                sync_target_volume_from_system()
                next_at = now + _get_audio_refresh_interval_seconds("volume_refresh_interval_ms")
        except Exception as e:
            log_error(f"Volume updater loop error: {e}")
            next_at = time.time() + _get_audio_refresh_interval_seconds("volume_refresh_interval_ms")
        _sleep_until(next_at, minimum_sleep=0.05, fallback_sleep=0.4)

def mute_updater_loop():
    next_at = 0.0
    while True:
        try:
            now = time.time()
            if now >= next_at:
                next_at = _update_mute_cache_tick(now)
        except Exception as e:
            log_error(f"Mute updater loop error: {e}")
            next_at = time.time() + _get_performance_interval_seconds("mute_refresh_interval_ms", 250, 100)
        _sleep_until(next_at, minimum_sleep=0.05, fallback_sleep=0.4)

def network_updater_loop():
    next_at = 0.0
    while True:
        try:
            now = time.time()
            if now >= next_at:
                next_at = _update_network_cache_tick(now)
        except Exception as e:
            log_error(f"Network updater loop error: {e}")
            next_at = time.time() + _get_network_refresh_interval_seconds()
        _sleep_until(next_at, minimum_sleep=0.05, fallback_sleep=0.8)

def uptime_updater_loop():
    next_at = 0.0
    while True:
        try:
            now = time.time()
            if now >= next_at:
                next_at = _update_uptime_cache_tick(now)
        except Exception as e:
            log_error(f"Uptime updater loop error: {e}")
            next_at = time.time() + _get_uptime_refresh_interval_seconds()
        _sleep_until(next_at, minimum_sleep=0.05, fallback_sleep=0.8)

def shift_updater_loop():
    next_at = 0.0
    while True:
        try:
            now = time.time()
            if now >= next_at:
                next_at = _update_shift_cache_tick(now)
        except Exception as e:
            log_error(f"Shift updater loop error: {e}")
            next_at = time.time() + max(60.0, float(_get_performance_int("shift_cache_check_interval_minutes", 30)) * 60.0)
        _sleep_until(next_at, minimum_sleep=0.1, fallback_sleep=1.0)

def update_system_cache(values=None, **kwargs):
    payload = {}
    if isinstance(values, dict):
        payload.update(values)
    payload.update(kwargs)
    if not payload:
        return
    payload.setdefault("last_update", time.time())
    changed = False
    with SENSOR_CACHE_LOCK:
        for key, value in payload.items():
            if key == "last_update":
                continue
            if SYSTEM_CACHE.get(key) != value:
                changed = True
                break
        SYSTEM_CACHE.update(payload)
    if changed:
        mark_system_cache_changed()

def get_system_snapshot():
    with SENSOR_CACHE_LOCK:
        return dict(SYSTEM_CACHE)

def get_system_cache_value(key, default=None):
    with SENSOR_CACHE_LOCK:
        return SYSTEM_CACHE.get(key, default)

def _load_shift_meta():
    try:
        with open(SHIFT_CACHE_META_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def _save_shift_meta(data):
    try:
        with open(SHIFT_CACHE_META_JSON, "w", encoding="utf-8") as f:
            json.dump(data or {}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log_error(f"Shift metadata write error: {e}")

def _normalize_shift_date_value(value):
    if value is None: return None
    if isinstance(value, datetime.datetime): return value.date()
    if isinstance(value, datetime.date): return value
    text = str(value).strip()
    if not text: return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try: return datetime.datetime.strptime(text, fmt).date()
        except Exception: pass
    return None

def _is_valid_shift_xlsx(path):
    try:
        if not path or (not os.path.exists(path)): return False
        with open(path, "rb") as f:
            sig = f.read(2)
        return sig == b"PK"
    except Exception: return False

def _download_shift_workbook_once_per_day(force=False):
    today_iso = datetime.date.today().isoformat()
    meta = _load_shift_meta()
    if (not force and meta.get("download_date") == today_iso and _is_valid_shift_xlsx(SHIFT_CACHE_XLSX)):
        return False
    
    url = _get_shift_share_url()
    if not url: return False
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = resp.read()
        if not data.startswith(b"PK"): return False
        with open(SHIFT_CACHE_XLSX, "wb") as f:
            f.write(data)
        _save_shift_meta({"download_date": today_iso, "url": url})
        SHIFT_RUNTIME["download_date"] = today_iso
        return True
    except Exception as e:
        log_error(f"Shift download error: {e}")
        return False

def _ensure_shift_schedule_loaded(force_download=False):
    downloaded = _download_shift_workbook_once_per_day(force=force_download)
    if downloaded or not SHIFT_RUNTIME.get("schedule_by_date"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(SHIFT_CACHE_XLSX, data_only=True, read_only=True)
            ws = wb[_get_shift_sheet_name()]
            emp = _get_shift_employee_name().lower()
            name_col = _get_shift_name_column()
            date_row = _get_shift_date_row()
            target_row = None
            for r in range(1, ws.max_row + 1):
                if str(ws.cell(r, name_col).value or "").lower() == emp:
                    target_row = r; break
            if not target_row: return False
            sched = {}
            for c in range(1, ws.max_column + 1):
                dt = _normalize_shift_date_value(ws.cell(date_row, c).value)
                if dt:
                    val = str(ws.cell(target_row, c).value or "").strip()
                    if val: sched[dt.isoformat()] = val
            SHIFT_RUNTIME["schedule_by_date"] = sched
            return True
        except Exception as e:
            log_error(f"Shift XLSX read error: {e}")
            return False
    return True



def _get_performance_interval_seconds(key, default_ms, min_ms):
    from panel_bootstrap import _get_performance_int
    ms = _get_performance_int(key, default_ms)
    return float(max(min_ms, ms)) / 1000.0

def _get_audio_refresh_interval_seconds(key='volume_refresh_interval_ms'):
    return _get_performance_interval_seconds(key, 250, 100)

def _get_network_refresh_interval_seconds():
    return _get_performance_interval_seconds('network_refresh_interval_ms', 5000, 1000)

def _get_uptime_refresh_interval_seconds():
    return _get_performance_interval_seconds('uptime_refresh_interval_ms', 5000, 1000)




def _update_shift_cache_tick(now_ts):
    now_dt = datetime.datetime.now()
    target_date = now_dt.date() if now_dt.hour < 9 else now_dt.date() + datetime.timedelta(days=1)
    _ensure_shift_schedule_loaded()
    shift_text = SHIFT_RUNTIME.get('schedule_by_date', {}).get(target_date.isoformat(), '--')
    if shift_text != '--':
        shift_text = re.split(r'[-—]', shift_text)[0].strip().replace('.', ':')
    update_system_cache(tomorrow_shift_text=shift_text, tomorrow_shift_subtitle=target_date.strftime('%d.%m'))
    from panel_bootstrap import _get_performance_int
    return now_ts + float(_get_performance_int('shift_cache_check_interval_minutes', 30)) * 60.0

def _order_tuya_devices_by_visible_keys(devices):
    from panel_bootstrap import _get_runtime_setting_cached
    keys = _get_runtime_setting_cached('tuya.visible_device_keys', [])
    if isinstance(keys, str): keys = [k.strip() for k in keys.split(',') if k.strip()]
    if not (isinstance(keys, list) and keys): return devices
    d_map = {str(d.get('key', '')).strip(): d for d in devices if d.get('key')}
    return [d_map[k] for k in keys if k in d_map]

def _build_public_status_payload_from_snapshot(snapshot):
    payload = {field: snapshot.get(field) for field in PUBLIC_STATUS_FIELDS}
    payload['tuya_devices'] = tuya_public_devices_payload(_order_tuya_devices_by_visible_keys(snapshot.get('tuya_devices') or []))
    payload['pc_plug_power_w'] = _get_pc_plug_power_w()
    return payload

def build_public_status_payload(cached=None):
    snapshot = cached if isinstance(cached, dict) else get_system_snapshot()
    if isinstance(cached, dict):
        return _build_public_status_payload_from_snapshot(snapshot)
    signature = (snapshot.get('last_update'), float(PC_PLUG_QUERY_CACHE.get('ts', 0)))
    with PUBLIC_STATUS_CACHE_LOCK:
        if PUBLIC_STATUS_CACHE.get('signature') == signature:
            return dict(PUBLIC_STATUS_CACHE['payload'])
    payload = _build_public_status_payload_from_snapshot(snapshot)
    with PUBLIC_STATUS_CACHE_LOCK:
        PUBLIC_STATUS_CACHE['signature'] = signature
        PUBLIC_STATUS_CACHE['payload'] = dict(payload)
    return payload

def _update_media_cache_tick(now):
    media_result = update_media_and_lyrics_cache()
    is_playing = bool(get_system_cache_value('media_is_playing', False))
    base_interval = _get_performance_interval_seconds('media_refresh_interval_ms', 500, 100)
    if is_playing: return now + base_interval
    return now + max(base_interval, 1.0, float(media_result.get('next_interval', 1.0)))

def _update_mute_cache_tick(now):
    from panel_audio_controls import get_system_mute_state_exact
    update_system_cache(is_muted=get_system_mute_state_exact())
    return now + _get_audio_refresh_interval_seconds('mute_refresh_interval_ms')

def _update_network_cache_tick(now):
    d, u = get_network_speed_mbps()
    update_system_cache(download_speed_mbps=d, upload_speed_mbps=u)
    return now + _get_network_refresh_interval_seconds()

def _update_uptime_cache_tick(now):
    update_system_cache(uptime=get_uptime_string())
    return now + _get_uptime_refresh_interval_seconds()
